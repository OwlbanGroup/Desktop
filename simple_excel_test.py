#!/usr/bin/env python3
"""
Simple test to create an Excel file and verify openpyxl functionality.
"""

import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    print("[PASS] Successfully imported openpyxl modules")
except ImportError as e:
    print(f"[FAIL] Failed to import openpyxl: {e}")
    sys.exit(1)

def create_simple_excel():
    """Create a simple Excel file to test functionality."""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"

        # Add some test data
        ws['A1'] = "Test Header"
        ws['B1'] = "Value"
        ws['A2'] = "Row 1"
        ws['B2'] = 123
        ws['A3'] = "Row 2"
        ws['B3'] = 456

        # Save the file
        filename = "simple_test.xlsx"
        wb.save(filename)

        if os.path.exists(filename):
            print(f"[PASS] Successfully created {filename}")
            file_size = os.path.getsize(filename)
            print(f"[PASS] File size: {file_size} bytes")
            return True
        else:
            print(f"[FAIL] File {filename} was not created")
            return False

    except Exception as e:
        print(f"[FAIL] Error creating Excel file: {e}")
        return False

if __name__ == "__main__":
    print("Testing Excel file creation...")
    success = create_simple_excel()
    if success:
        print("[SUCCESS] Excel file creation test passed!")
        sys.exit(0)
    else:
        print("[ERROR] Excel file creation test failed!")
        sys.exit(1)
